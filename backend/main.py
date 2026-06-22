import os
import re
import io
import json
import base64
import hashlib
import shutil
import datetime
import anthropic
from contextlib import asynccontextmanager
from typing import Optional, List
from fastapi import FastAPI, File, UploadFile, Depends, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from dotenv import load_dotenv
from jose import JWTError, jwt
import bcrypt as _bcrypt
from sqlalchemy import select, func

# Load environmental variables from .env
load_dotenv()

from database import (
    init_db, get_db, User, DailyFoodLog, BrandPreference, IngredientCache,
    FrequentMeal, APIKey, update_user_streak, get_brand_preferences,
    set_brand_preference, delete_brand_preference,
    get_all_api_keys, save_api_key, delete_api_key,
    get_app_setting, set_app_setting,
    create_user, get_user_by_username, get_user_by_id,
)
from sqlalchemy.ext.asyncio import AsyncSession
from stt_worker import transcribe_audio
from graph_engine import compiled_graph

# ── Environment detection ─────────────────────────────────────────────────────
IS_PRODUCTION = os.environ.get("ENVIRONMENT", "development").strip().lower() == "production"

# ── JWT / Auth configuration ──────────────────────────────────────────────────
# Set JWT_SECRET to a long random string in your hosting provider's env vars.
# The fallback is fine for local development but MUST be overridden in production.
SECRET_KEY = os.environ.get("JWT_SECRET", "fitvoice-dev-secret-change-me-in-production")
ALGORITHM  = "HS256"
TOKEN_EXPIRE_DAYS = 7

http_bearer = HTTPBearer(auto_error=False)

# Use bcrypt directly — passlib 1.7.4 is incompatible with bcrypt 4.x because
# passlib's internal wrap-bug detection uses a 73-byte test string which newer
# bcrypt now rejects. Calling bcrypt directly avoids that internal check.
def hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode("utf-8")[:72], _bcrypt.gensalt()).decode()

def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode("utf-8")[:72], hashed.encode())

def create_access_token(user_id: int) -> str:
    expire = datetime.datetime.utcnow() + datetime.timedelta(days=TOKEN_EXPIRE_DAYS)
    return jwt.encode({"sub": str(user_id), "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(http_bearer),
    db: AsyncSession = Depends(get_db),
) -> User:
    """FastAPI dependency — validate JWT and return the requesting User."""
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated. Please log in.")
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: Optional[str] = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token payload.")
    except JWTError:
        raise HTTPException(status_code=401, detail="Token is invalid or has expired. Please log in again.")

    user = await get_user_by_id(db, int(user_id))
    if not user:
        raise HTTPException(status_code=401, detail="User account no longer exists.")
    return user

# ── Supported API providers ───────────────────────────────────────────────────
SUPPORTED_PROVIDERS = {
    "anthropic": {"label": "Anthropic Claude",  "env_key": "ANTHROPIC_API_KEY",  "description": "Used for ingredient extraction, macro resolution & label vision (required)"},
    "openai":    {"label": "OpenAI",             "env_key": "OPENAI_API_KEY",     "description": "Alternative LLM for extraction and resolution"},
    "groq":      {"label": "Groq",               "env_key": "GROQ_API_KEY",       "description": "Ultra-fast Whisper API for speech-to-text"},
    "gemini":    {"label": "Google Gemini",      "env_key": "GEMINI_API_KEY",     "description": "Google's multimodal LLM alternative"},
    "tavily":    {"label": "Tavily Search",      "env_key": "TAVILY_API_KEY",     "description": "Web search fallback for unknown ingredients"},
}

# ── App lifespan (DB init + key hot-loading) ──────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    os.makedirs("temp_audio", exist_ok=True)
    from database import AsyncSessionLocal
    async with AsyncSessionLocal() as session:
        keys = await get_all_api_keys(session)
        for k in keys:
            if k.provider in SUPPORTED_PROVIDERS:
                os.environ[SUPPORTED_PROVIDERS[k.provider]["env_key"]] = k.api_key
        if IS_PRODUCTION:
            os.environ["STT_MODE"] = "cloud"
        else:
            saved_mode = await get_app_setting(session, "stt_mode", default="auto")
            os.environ["STT_MODE"] = saved_mode
    yield
    if os.path.exists("temp_audio"):
        shutil.rmtree("temp_audio")

app = FastAPI(
    title="FitVoice API",
    description="Voice-driven meal tracker — macro resolution with LangGraph + Claude",
    version="2.0.0",
    lifespan=lifespan,
)

# ── CORS ──────────────────────────────────────────────────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "https://fit-ai-black-one.vercel.app",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Pydantic schemas ──────────────────────────────────────────────────────────

class SignupSchema(BaseModel):
    username: str
    password: str
    name: str

class LoginSchema(BaseModel):
    username: str
    password: str

class UserUpdateSchema(BaseModel):
    name: str
    target_calories: float
    target_protein: float
    target_carbs: float
    target_fat: float

class BrandPreferenceSchema(BaseModel):
    ingredient_name: str
    preferred_brand: str

class APIKeySchema(BaseModel):
    provider: str
    api_key: str

class STTModeSchema(BaseModel):
    mode: str  # "auto" | "cloud" | "local"

VALID_STT_MODES = {"auto", "cloud", "local"}

# ─────────────────────────────────────────────────────────────────────────────
# AUTH ENDPOINTS
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health_check():
    return {"status": "healthy", "time": datetime.datetime.now().isoformat()}

@app.post("/api/auth/signup")
async def signup(payload: SignupSchema, db: AsyncSession = Depends(get_db)):
    """Register a new user. Returns a JWT token on success."""
    if len(payload.username.strip()) < 3:
        raise HTTPException(status_code=400, detail="Username must be at least 3 characters.")
    if len(payload.password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters.")
    if len(payload.name.strip()) < 1:
        raise HTTPException(status_code=400, detail="Name is required.")

    existing = await get_user_by_username(db, payload.username)
    if existing:
        raise HTTPException(status_code=400, detail="Username already taken. Please choose another.")

    user = await create_user(
        db,
        username=payload.username,
        name=payload.name.strip(),
        password_hash=hash_password(payload.password),
    )
    token = create_access_token(user.id)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user.id, "name": user.name, "username": user.username},
    }

@app.post("/api/auth/login")
async def login(payload: LoginSchema, db: AsyncSession = Depends(get_db)):
    """Authenticate and return a JWT token."""
    user = await get_user_by_username(db, payload.username)
    if not user or not user.password_hash or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid username or password.")
    token = create_access_token(user.id)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user.id, "name": user.name, "username": user.username},
    }

# ─────────────────────────────────────────────────────────────────────────────
# USER PROFILE (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/user")
async def get_user_profile(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "name": current_user.name,
        "username": current_user.username,
        "target_calories": current_user.target_calories,
        "target_protein": current_user.target_protein,
        "target_carbs": current_user.target_carbs,
        "target_fat": current_user.target_fat,
        "current_streak": current_user.current_streak,
        "last_active_date": current_user.last_active_date.isoformat() if current_user.last_active_date else None,
    }

@app.post("/api/user")
async def update_user_profile(
    payload: UserUpdateSchema,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    current_user.name = payload.name
    current_user.target_calories = payload.target_calories
    current_user.target_protein = payload.target_protein
    current_user.target_carbs = payload.target_carbs
    current_user.target_fat = payload.target_fat
    await db.commit()
    return {"status": "success", "user": {
        "name": current_user.name,
        "target_calories": current_user.target_calories,
        "target_protein": current_user.target_protein,
        "target_carbs": current_user.target_carbs,
        "target_fat": current_user.target_fat,
    }}

# ─────────────────────────────────────────────────────────────────────────────
# DASHBOARD (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/dashboard")
async def get_dashboard(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    today = datetime.date.today()
    logs_result = await db.execute(
        select(DailyFoodLog)
        .where(DailyFoodLog.user_id == current_user.id, DailyFoodLog.date == today)
        .order_by(DailyFoodLog.id.desc())
    )
    logs = logs_result.scalars().all()

    today_calories = today_protein = today_carbs = today_fat = 0.0
    meals_list = []
    for log in logs:
        macros = log.computed_macros.get("total_meal_macros", {})
        today_calories += macros.get("calories", 0.0)
        today_protein  += macros.get("protein", 0.0)
        today_carbs    += macros.get("carbs", 0.0)
        today_fat      += macros.get("fat", 0.0)
        meals_list.append({
            "id": log.id,
            "raw_transcript": log.raw_transcript,
            "date": log.date.isoformat(),
            "macros": macros,
            "ingredients": log.computed_macros.get("resolved_ingredients", []),
        })

    return {
        "user": {
            "name": current_user.name,
            "username": current_user.username,
            "current_streak": current_user.current_streak,
            "target_calories": current_user.target_calories,
            "target_protein": current_user.target_protein,
            "target_carbs": current_user.target_carbs,
            "target_fat": current_user.target_fat,
        },
        "totals": {
            "calories": round(today_calories, 1),
            "protein": round(today_protein, 1),
            "carbs": round(today_carbs, 1),
            "fat": round(today_fat, 1),
        },
        "meals": meals_list,
    }

# ─────────────────────────────────────────────────────────────────────────────
# TRANSCRIBE (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/transcribe")
async def transcribe_only(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    temp_file_path = f"temp_audio/uploaded_{datetime.datetime.now().timestamp()}_{file.filename}"
    try:
        with open(temp_file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        transcript = transcribe_audio(temp_file_path)
        return {"status": "success", "transcript": transcript}
    except Exception as e:
        raise HTTPException(status_code=500, detail="Failed to transcribe audio")
    finally:
        if os.path.exists(temp_file_path):
            try: os.remove(temp_file_path)
            except Exception: pass

# ─────────────────────────────────────────────────────────────────────────────
# FREQUENT MEAL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def meal_fingerprint(resolved_ingredients: list) -> str:
    """MD5 of sorted, lowercased ingredient names — stable identity for a meal."""
    names = sorted(
        (ing.get("name", "") or "").strip().lower()
        for ing in resolved_ingredients
        if ing.get("name")
    )
    return hashlib.md5("|".join(names).encode()).hexdigest()

def meal_display_name(resolved_ingredients: list) -> str:
    """Human-readable name: first 3 ingredient names joined by ' + '."""
    names = [ing.get("name", "").strip().title() for ing in resolved_ingredients if ing.get("name")]
    if not names:
        return "Meal"
    if len(names) <= 3:
        return " + ".join(names)
    return " + ".join(names[:3]) + f" +{len(names) - 3} more"

async def _upsert_frequent_meal(
    db: AsyncSession,
    user_id: int,
    resolved_ingredients: list,
    total_meal_macros: dict,
) -> None:
    """Called after every successful meal log. Creates or increments the frequent meal record."""
    fp = meal_fingerprint(resolved_ingredients)
    existing_res = await db.execute(
        select(FrequentMeal).where(
            FrequentMeal.user_id == user_id,
            FrequentMeal.meal_fingerprint == fp,
        )
    )
    fm = existing_res.scalar_one_or_none()
    today = datetime.date.today()
    if fm:
        fm.log_count  += 1
        fm.last_logged = today
        fm.macros      = total_meal_macros       # keep latest macro snapshot
        fm.ingredients = resolved_ingredients    # keep latest portion sizes
    else:
        db.add(FrequentMeal(
            user_id=user_id,
            meal_fingerprint=fp,
            display_name=meal_display_name(resolved_ingredients),
            ingredients=resolved_ingredients,
            macros=total_meal_macros,
            log_count=1,
            last_logged=today,
        ))
    # commit is handled by the caller after adding the food log

# ─────────────────────────────────────────────────────────────────────────────
# TRACK MEAL (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/api/track-meal")
async def track_meal(
    file: Optional[UploadFile] = File(None),
    text: Optional[str] = Form(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    transcript = ""
    if file:
        temp_file_path = f"temp_audio/uploaded_{datetime.datetime.now().timestamp()}_{file.filename}"
        try:
            with open(temp_file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            transcript = transcribe_audio(temp_file_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail="Failed to process or transcribe audio")
        finally:
            if os.path.exists(temp_file_path):
                try: os.remove(temp_file_path)
                except Exception: pass
    elif text:
        transcript = text.strip()
    else:
        raise HTTPException(status_code=400, detail="Either audio file or text is required.")

    if not transcript:
        raise HTTPException(status_code=400, detail="Could not capture speech. Please try again.")

    try:
        final_state = compiled_graph.invoke({"raw_text": transcript})
    except Exception as ge:
        raise HTTPException(status_code=500, detail=f"Macro resolution failed: {ge}")

    resolved_ingredients = final_state.get("resolved_ingredients", [])
    total_meal_macros    = final_state.get("total_meal_macros", {})

    streak = await update_user_streak(db, user_id=current_user.id)

    food_log = DailyFoodLog(
        user_id=current_user.id,
        date=datetime.date.today(),
        raw_transcript=transcript,
        computed_macros={"resolved_ingredients": resolved_ingredients, "total_meal_macros": total_meal_macros},
    )
    db.add(food_log)

    # Auto-upsert frequent meal record (creates or increments log_count)
    await _upsert_frequent_meal(db, current_user.id, resolved_ingredients, total_meal_macros)

    await db.commit()

    return {
        "status": "success",
        "transcript": transcript,
        "streak": streak,
        "ingredients": resolved_ingredients,
        "macros": total_meal_macros,
    }

# ─────────────────────────────────────────────────────────────────────────────
# PROGRESS — 90-day heatmap data (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/progress")
async def get_progress(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Returns 90 days of daily calorie summaries (for the GitHub-style heatmap)
    plus all-time stats (streaks, totals).
    """
    today           = datetime.date.today()
    ninety_days_ago = today - datetime.timedelta(days=89)

    # Fetch last 90 days of logs for this user
    recent = await db.execute(
        select(DailyFoodLog)
        .where(DailyFoodLog.user_id == current_user.id, DailyFoodLog.date >= ninety_days_ago)
        .order_by(DailyFoodLog.date)
    )
    recent_logs = recent.scalars().all()

    # Group into a date → macro totals dict
    daily_map: dict = {}
    for log in recent_logs:
        ds = log.date.isoformat()
        if ds not in daily_map:
            daily_map[ds] = {"calories": 0, "protein": 0, "carbs": 0, "fat": 0, "meal_count": 0}
        m = log.computed_macros.get("total_meal_macros", {})
        daily_map[ds]["calories"]   += m.get("calories", 0)
        daily_map[ds]["protein"]    += m.get("protein", 0)
        daily_map[ds]["carbs"]      += m.get("carbs", 0)
        daily_map[ds]["fat"]        += m.get("fat", 0)
        daily_map[ds]["meal_count"] += 1

    # Build 90-cell array (oldest first)
    summaries = []
    for i in range(90):
        d  = ninety_days_ago + datetime.timedelta(days=i)
        ds = d.isoformat()
        if ds in daily_map:
            cal   = round(daily_map[ds]["calories"], 1)
            ratio = cal / current_user.target_calories if current_user.target_calories else 0
            if ratio >= 0.9 and ratio <= 1.15:
                status = "met"        # sweet spot — green
            elif ratio > 1.15:
                status = "over"       # above target — orange/red
            elif ratio >= 0.5:
                status = "under"      # logged but under target — muted green
            else:
                status = "minimal"    # logged almost nothing — very muted
            summaries.append({
                "date": ds,
                "calories": cal,
                "protein": round(daily_map[ds]["protein"], 1),
                "carbs":   round(daily_map[ds]["carbs"], 1),
                "fat":     round(daily_map[ds]["fat"], 1),
                "meal_count": daily_map[ds]["meal_count"],
                "target_calories": current_user.target_calories,
                "status": status,
            })
        else:
            summaries.append({
                "date": ds, "calories": 0, "protein": 0, "carbs": 0, "fat": 0,
                "meal_count": 0, "target_calories": current_user.target_calories,
                "status": "empty",
            })

    # All-time stats for this user
    all_logs_res = await db.execute(
        select(DailyFoodLog).where(DailyFoodLog.user_id == current_user.id)
    )
    all_logs = all_logs_res.scalars().all()

    unique_dates = sorted(set(log.date for log in all_logs))
    best_streak = curr = 0
    prev_d = None
    for d in unique_dates:
        curr = (curr + 1) if prev_d and (d - prev_d).days == 1 else 1
        best_streak = max(best_streak, curr)
        prev_d = d

    total_count_res = await db.execute(
        select(func.count(DailyFoodLog.id)).where(DailyFoodLog.user_id == current_user.id)
    )
    total_meals = total_count_res.scalar() or 0

    return {
        "summaries": summaries,
        "stats": {
            "current_streak": current_user.current_streak,
            "best_streak": best_streak,
            "total_days_logged": len(unique_dates),
            "total_meals": total_meals,
        },
        "target_calories": current_user.target_calories,
    }

# ─────────────────────────────────────────────────────────────────────────────
# HISTORY — paginated past meals (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/history")
async def get_history(
    page: int = 1,
    per_page: int = 20,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    offset = (page - 1) * per_page
    logs_res = await db.execute(
        select(DailyFoodLog)
        .where(DailyFoodLog.user_id == current_user.id)
        .order_by(DailyFoodLog.date.desc(), DailyFoodLog.id.desc())
        .offset(offset)
        .limit(per_page)
    )
    logs = logs_res.scalars().all()

    total_res = await db.execute(
        select(func.count(DailyFoodLog.id)).where(DailyFoodLog.user_id == current_user.id)
    )
    total = total_res.scalar() or 0

    return {
        "meals": [
            {
                "id": log.id,
                "date": log.date.isoformat(),
                "raw_transcript": log.raw_transcript,
                "macros": log.computed_macros.get("total_meal_macros", {}),
                "ingredients": log.computed_macros.get("resolved_ingredients", []),
            }
            for log in logs
        ],
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page),
    }

# ─────────────────────────────────────────────────────────────────────────────
# DELETE MEAL (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.delete("/api/meals/{meal_id}")
async def delete_meal(
    meal_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Delete a specific meal log. Users can only delete their own meals."""
    result = await db.execute(
        select(DailyFoodLog).where(
            DailyFoodLog.id == meal_id,
            DailyFoodLog.user_id == current_user.id,   # ownership check
        )
    )
    meal = result.scalar_one_or_none()
    if not meal:
        raise HTTPException(status_code=404, detail="Meal not found or does not belong to you.")
    await db.delete(meal)
    await db.commit()
    return {"status": "deleted", "meal_id": meal_id}

# ─────────────────────────────────────────────────────────────────────────────
# FREQUENT MEALS (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

class PortionAdjustSchema(BaseModel):
    """Optional per-ingredient gram overrides for quick-log portion editor."""
    portions: Optional[dict] = None   # { "ingredient_name": new_grams }

@app.get("/api/frequent-meals")
async def list_frequent_meals(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return meals logged at least twice, sorted by log_count desc."""
    res = await db.execute(
        select(FrequentMeal)
        .where(FrequentMeal.user_id == current_user.id, FrequentMeal.log_count >= 2)
        .order_by(FrequentMeal.log_count.desc(), FrequentMeal.last_logged.desc())
        .limit(20)
    )
    meals = res.scalars().all()
    return [
        {
            "id": m.id,
            "display_name": m.display_name,
            "ingredients": m.ingredients,
            "macros": m.macros,
            "log_count": m.log_count,
            "last_logged": m.last_logged.isoformat(),
        }
        for m in meals
    ]

@app.post("/api/frequent-meals/{meal_id}/log")
async def quick_log_frequent_meal(
    meal_id: int,
    payload: PortionAdjustSchema,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Quick-log a frequent meal. Optionally pass `portions` to scale ingredient grams.
    Portions dict: { "chicken breast": 200 }  (key = ingredient name, value = new grams)
    """
    res = await db.execute(
        select(FrequentMeal).where(
            FrequentMeal.id == meal_id,
            FrequentMeal.user_id == current_user.id,
        )
    )
    fm = res.scalar_one_or_none()
    if not fm:
        raise HTTPException(status_code=404, detail="Frequent meal not found.")

    ingredients = list(fm.ingredients)   # make a mutable copy

    # Apply portion overrides if provided
    if payload.portions:
        for ing in ingredients:
            name = (ing.get("name") or "").lower()
            if name in {k.lower() for k in payload.portions}:
                # Find the matching key (case-insensitive)
                override_key = next(k for k in payload.portions if k.lower() == name)
                new_grams    = float(payload.portions[override_key])
                old_grams    = float(ing.get("grams", 100))
                ratio        = new_grams / old_grams if old_grams else 1.0
                # Scale every macro proportionally
                for macro in ("calories", "protein", "carbs", "fat"):
                    if macro in ing:
                        ing[macro] = round(float(ing[macro]) * ratio, 1)
                ing["grams"] = new_grams

    # Recompute total macros after portion adjustment
    total_cal = sum(float(i.get("calories", 0)) for i in ingredients)
    total_pro = sum(float(i.get("protein",  0)) for i in ingredients)
    total_crb = sum(float(i.get("carbs",    0)) for i in ingredients)
    total_fat = sum(float(i.get("fat",      0)) for i in ingredients)
    total_meal_macros = {
        "calories": round(total_cal, 1),
        "protein":  round(total_pro, 1),
        "carbs":    round(total_crb, 1),
        "fat":      round(total_fat, 1),
    }

    transcript = f"Quick-logged: {fm.display_name}"

    food_log = DailyFoodLog(
        user_id=current_user.id,
        date=datetime.date.today(),
        raw_transcript=transcript,
        computed_macros={"resolved_ingredients": ingredients, "total_meal_macros": total_meal_macros},
    )
    db.add(food_log)

    # Also keep the frequent-meal record updated
    fm.log_count  += 1
    fm.last_logged = datetime.date.today()

    streak = await update_user_streak(db, user_id=current_user.id)
    await db.commit()

    return {
        "status": "success",
        "streak": streak,
        "ingredients": ingredients,
        "macros": total_meal_macros,
        "display_name": fm.display_name,
    }

@app.delete("/api/frequent-meals/{meal_id}")
async def delete_frequent_meal(
    meal_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(
        select(FrequentMeal).where(
            FrequentMeal.id == meal_id,
            FrequentMeal.user_id == current_user.id,
        )
    )
    fm = res.scalar_one_or_none()
    if not fm:
        raise HTTPException(status_code=404, detail="Frequent meal not found.")
    await db.delete(fm)
    await db.commit()
    return {"status": "deleted", "meal_id": meal_id}

# ─────────────────────────────────────────────────────────────────────────────
# SMART MACRO SUGGESTIONS — Claude-powered (auth-protected)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/suggest-meals")
async def suggest_meals(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Computes today's remaining macros and asks Claude to suggest 3 meals that
    best fill the gap — preferring the user's frequent meals where possible.
    Returns a ranked list with fit_score and fit_reason.
    """
    # 1. Today's consumed macros
    today = datetime.date.today()
    logs_res = await db.execute(
        select(DailyFoodLog).where(
            DailyFoodLog.user_id == current_user.id,
            DailyFoodLog.date == today,
        )
    )
    logs = logs_res.scalars().all()

    consumed = {"calories": 0.0, "protein": 0.0, "carbs": 0.0, "fat": 0.0}
    for log in logs:
        m = log.computed_macros.get("total_meal_macros", {})
        for k in consumed:
            consumed[k] += float(m.get(k, 0))

    remaining = {
        "calories": round(current_user.target_calories - consumed["calories"], 1),
        "protein":  round(current_user.target_protein  - consumed["protein"],  1),
        "carbs":    round(current_user.target_carbs    - consumed["carbs"],    1),
        "fat":      round(current_user.target_fat      - consumed["fat"],      1),
    }

    # 2. Frequent meals (top 10 by count)
    fm_res = await db.execute(
        select(FrequentMeal)
        .where(FrequentMeal.user_id == current_user.id, FrequentMeal.log_count >= 2)
        .order_by(FrequentMeal.log_count.desc())
        .limit(10)
    )
    frequent = fm_res.scalars().all()
    frequent_summaries = [
        {
            "id": f.id,
            "name": f.display_name,
            "macros": f.macros,
            "log_count": f.log_count,
        }
        for f in frequent
    ]

    # 3. Ask Claude
    prompt = f"""You are a nutrition coach for FitVoice, a macro-tracking app.

The user's REMAINING macros for today are:
- Calories: {remaining['calories']} kcal
- Protein:  {remaining['protein']} g
- Carbs:    {remaining['carbs']} g
- Fat:      {remaining['fat']} g

Their frequently logged meals (prefer these when they fit):
{json.dumps(frequent_summaries, indent=2)}

Suggest exactly 3 meals that best fill the remaining macros.
Rules:
- Prefer frequent meals from the list above when they fit well.
- If a frequent meal is used, set "is_frequent": true and "frequent_meal_id" to its id.
- Otherwise set "is_frequent": false and "frequent_meal_id": null.
- fit_score: integer 1-10 (10 = perfect macro fit).
- fit_reason: one short sentence explaining why this meal fits.
- estimated_macros: your best estimate for this meal's macros.
- description: brief description (e.g. "250g grilled chicken + 150g rice").

Reply with ONLY a valid JSON array of 3 objects, no markdown, no explanation:
[
  {{
    "name": "...",
    "description": "...",
    "is_frequent": true,
    "frequent_meal_id": 12,
    "estimated_macros": {{"calories": 0, "protein": 0, "carbs": 0, "fat": 0}},
    "fit_score": 9,
    "fit_reason": "..."
  }},
  ...
]"""

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        # Strip potential markdown fences
        raw = re.sub(r"^```[a-z]*\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        suggestions: List[dict] = json.loads(raw)
    except Exception as e:
        # Return a graceful fallback so the UI doesn't break
        suggestions = [
            {
                "name": "High-protein meal",
                "description": f"A meal targeting ~{remaining['protein']}g protein",
                "is_frequent": False,
                "frequent_meal_id": None,
                "estimated_macros": remaining,
                "fit_score": 5,
                "fit_reason": "Suggestion service temporarily unavailable.",
            }
        ]

    return {
        "remaining": remaining,
        "suggestions": suggestions,
    }

# ─────────────────────────────────────────────────────────────────────────────
# BRAND PREFERENCES (global — no auth required)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/brand-preferences")
async def list_brand_preferences(db: AsyncSession = Depends(get_db)):
    prefs = await get_brand_preferences(db)
    return [{"ingredient_name": p.ingredient_name, "preferred_brand": p.preferred_brand} for p in prefs]

@app.post("/api/brand-preferences")
async def upsert_brand_preference(payload: BrandPreferenceSchema, db: AsyncSession = Depends(get_db)):
    pref = await set_brand_preference(db, payload.ingredient_name, payload.preferred_brand)
    return {"status": "success", "ingredient_name": pref.ingredient_name, "preferred_brand": pref.preferred_brand}

@app.delete("/api/brand-preferences/{ingredient_name}")
async def remove_brand_preference(ingredient_name: str, db: AsyncSession = Depends(get_db)):
    deleted = await delete_brand_preference(db, ingredient_name)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"No preference found for '{ingredient_name}'")
    return {"status": "deleted", "ingredient_name": ingredient_name}

@app.get("/api/brand-preferences/export")
async def export_brand_preferences(db: AsyncSession = Depends(get_db)):
    """
    Export all brand preferences + their cached nutrition data as a styled .xlsx file.
    The downloaded file can be re-uploaded via /api/brand-preferences/import to bulk-restore
    preferences on a fresh deployment or share them with another device.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    # ── Fetch all preferences + cache entries ────────────────────────────────
    prefs  = await get_brand_preferences(db)
    caches_res = await db.execute(select(IngredientCache))
    caches = caches_res.scalars().all()

    # Build lookup: (ingredient_name, brand) → cache row
    cache_map: dict = {}
    for c in caches:
        key = (c.name.lower(), (c.brand or "").lower())
        cache_map[key] = c

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Brand Preferences"

    # ── Colour palette
    HEADER_BG   = "4F46E5"   # indigo-600
    HEADER_FG   = "FFFFFF"
    ALT_ROW_BG  = "F1F0FF"   # very light indigo tint
    BORDER_CLR  = "C7D2FE"   # indigo-200
    NOTE_BG     = "EEF2FF"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Instructions row ─────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = (
        "FitVoice — Brand Preferences Export  |  "
        "Columns A–B are required for import. "
        "Edit nutritional values (C–F) and re-upload to update the ingredient cache."
    )
    ws["A1"].font      = Font(name="Arial", size=9, italic=True, color="6366F1")
    ws["A1"].fill      = PatternFill("solid", fgColor=NOTE_BG)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        ("Ingredient",       "A", 22),
        ("Brand",            "B", 22),
        ("Calories / 100g",  "C", 17),
        ("Protein / 100g",   "D", 17),
        ("Carbs / 100g",     "E", 17),
        ("Fat / 100g",       "F", 17),
    ]

    for col_idx, (label, col_letter, width) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font      = Font(name="Arial", size=10, bold=True, color=HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[2].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, pref in enumerate(prefs, start=3):
        name_lower  = pref.ingredient_name.lower()
        brand_lower = pref.preferred_brand.lower()
        cache_entry = cache_map.get((name_lower, brand_lower)) or cache_map.get((name_lower, ""))

        row_fill = PatternFill("solid", fgColor=ALT_ROW_BG) if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        data = [
            pref.ingredient_name,
            pref.preferred_brand,
            round(cache_entry.calories_per_100g, 1) if cache_entry else "",
            round(cache_entry.protein_per_100g,  1) if cache_entry else "",
            round(cache_entry.carbs_per_100g,    1) if cache_entry else "",
            round(cache_entry.fat_per_100g,      1) if cache_entry else "",
        ]

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = row_fill
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")
            cell.border    = border

        ws.row_dimensions[row_idx].height = 18

    # ── Freeze header rows + auto-filter ─────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:F{max(2, 1 + len(prefs))}"

    # ── Stream back as .xlsx ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"fitvoice_brand_preferences_{datetime.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/brand-preferences/import")
async def import_brand_preferences(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Bulk-import brand preferences from an .xlsx file (same format as the export).
    Each row must have at minimum: Ingredient (col A) and Brand (col B).
    Nutritional columns C–F, if filled, update the ingredient cache as well.
    Returns a summary of rows imported vs. skipped.
    """
    from openpyxl import load_workbook

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Could not parse file. Please upload a valid .xlsx file.")

    ws = wb.active
    imported = skipped = 0
    errors: list[str] = []

    for row in ws.iter_rows(min_row=3, values_only=True):  # row 1=note, row 2=header
        # Skip fully empty rows
        if not any(row):
            continue

        ingredient = str(row[0]).strip() if row[0] is not None else ""
        brand      = str(row[1]).strip() if row[1] is not None else ""

        if not ingredient or not brand:
            skipped += 1
            continue

        name_lower  = ingredient.lower()
        brand_lower = brand.lower()

        # Save brand preference
        await set_brand_preference(db, ingredient, brand)

        # If nutritional columns are provided, upsert ingredient cache
        try:
            cal  = float(row[2]) if row[2] not in (None, "") else None
            prot = float(row[3]) if row[3] not in (None, "") else None
            carb = float(row[4]) if row[4] not in (None, "") else None
            fat  = float(row[5]) if row[5] not in (None, "") else None

            if all(v is not None for v in [cal, prot, carb, fat]):
                existing = await db.execute(
                    select(IngredientCache).where(
                        IngredientCache.name  == name_lower,
                        IngredientCache.brand == brand_lower,
                    )
                )
                entry = existing.scalar_one_or_none()
                if entry:
                    entry.calories_per_100g = cal
                    entry.protein_per_100g  = prot
                    entry.carbs_per_100g    = carb
                    entry.fat_per_100g      = fat
                else:
                    db.add(IngredientCache(
                        name=name_lower, brand=brand_lower,
                        calories_per_100g=cal, protein_per_100g=prot,
                        carbs_per_100g=carb, fat_per_100g=fat,
                    ))
        except (ValueError, TypeError) as e:
            errors.append(f"Row '{ingredient}/{brand}': invalid nutrition value — skipped cache update.")

        imported += 1

    await db.commit()
    return {
        "status": "success",
        "imported": imported,
        "skipped": skipped,
        "warnings": errors,
        "message": f"{imported} preference(s) imported{f', {skipped} skipped' if skipped else ''}.",
    }


@app.post("/api/brand-preferences/label")
async def set_brand_from_label(
    ingredient_name: str = Form(...),
    preferred_brand: str = Form(...),
    image: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Extract macros from a nutrition label photo via Claude vision and persist them."""
    image_data = await image.read()
    image_b64  = base64.b64encode(image_data).decode()
    media_type = image.content_type or "image/jpeg"
    client = anthropic.Anthropic()

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=512,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}},
                {"type": "text", "text": (
                    f"This is a nutrition label for '{preferred_brand} {ingredient_name}'. "
                    "Extract the nutritional values and normalise them to per 100g. "
                    "Reply with ONLY a raw JSON object with these four numeric keys: "
                    "calories_per_100g, protein_per_100g, carbs_per_100g, fat_per_100g."
                )},
            ]}],
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Claude vision call failed: {e}")

    raw_text   = response.content[0].text.strip()
    json_match = re.search(r'\{.*?\}', raw_text, re.DOTALL)
    if not json_match:
        raise HTTPException(status_code=422, detail="Could not parse nutrition data from label.")

    try:
        macros   = json.loads(json_match.group())
        required = {"calories_per_100g", "protein_per_100g", "carbs_per_100g", "fat_per_100g"}
        if not required.issubset(macros.keys()):
            raise ValueError("Missing keys")
        macros = {k: float(macros[k]) for k in required}
    except Exception:
        raise HTTPException(status_code=422, detail="Label extraction returned incomplete data.")

    name_lower  = ingredient_name.strip().lower()
    brand_lower = preferred_brand.strip().lower()

    from sqlalchemy import select as sql_select
    result = await db.execute(
        sql_select(IngredientCache).where(IngredientCache.name == name_lower, IngredientCache.brand == brand_lower)
    )
    cache_entry = result.scalar_one_or_none()
    if cache_entry:
        cache_entry.calories_per_100g = macros["calories_per_100g"]
        cache_entry.protein_per_100g  = macros["protein_per_100g"]
        cache_entry.carbs_per_100g    = macros["carbs_per_100g"]
        cache_entry.fat_per_100g      = macros["fat_per_100g"]
    else:
        db.add(IngredientCache(name=name_lower, brand=brand_lower, **macros))

    await set_brand_preference(db, ingredient_name, preferred_brand)
    await db.commit()
    return {"status": "success", "ingredient_name": name_lower, "preferred_brand": brand_lower, "macros": macros, "source": "label_image"}

# ─────────────────────────────────────────────────────────────────────────────
# API KEY MANAGEMENT (global — no auth)
# ─────────────────────────────────────────────────────────────────────────────

def mask_key(key: str) -> str:
    if len(key) <= 8:
        return "••••••••"
    return key[:4] + "••••••••" + key[-4:]

@app.get("/api/keys")
async def list_api_keys(db: AsyncSession = Depends(get_db)):
    saved  = {k.provider: k.api_key for k in await get_all_api_keys(db)}
    result = []
    for provider, meta in SUPPORTED_PROVIDERS.items():
        key = saved.get(provider) or os.environ.get(meta["env_key"], "")
        result.append({
            "provider": provider, "label": meta["label"],
            "description": meta["description"], "env_key": meta["env_key"],
            "is_set": bool(key), "masked_key": mask_key(key) if key else "",
        })
    return result

@app.post("/api/keys")
async def upsert_api_key(payload: APIKeySchema, db: AsyncSession = Depends(get_db)):
    if payload.provider not in SUPPORTED_PROVIDERS:
        raise HTTPException(status_code=400, detail=f"Unknown provider '{payload.provider}'")
    await save_api_key(db, payload.provider, payload.api_key)
    os.environ[SUPPORTED_PROVIDERS[payload.provider]["env_key"]] = payload.api_key
    return {"status": "saved", "provider": payload.provider, "masked_key": mask_key(payload.api_key)}

@app.delete("/api/keys/{provider}")
async def remove_api_key(provider: str, db: AsyncSession = Depends(get_db)):
    if provider not in SUPPORTED_PROVIDERS:
        raise HTTPException(status_code=400, detail=f"Unknown provider '{provider}'")
    deleted = await delete_api_key(db, provider)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"No saved key for '{provider}'")
    os.environ.pop(SUPPORTED_PROVIDERS[provider]["env_key"], None)
    return {"status": "deleted", "provider": provider}

# ─────────────────────────────────────────────────────────────────────────────
# STT MODE SETTINGS (dev-only toggle, no auth required)
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/api/stt-settings")
async def get_stt_settings(db: AsyncSession = Depends(get_db)):
    return {
        "allow_local_choice": not IS_PRODUCTION,
        "current_mode": os.environ.get("STT_MODE", "auto"),
        "is_production": IS_PRODUCTION,
        "modes": [
            {"value": "auto",  "label": "Auto (Cloud then Local)",       "description": "Uses Groq if key is set, falls back to local model."},
            {"value": "cloud", "label": "Cloud Only (Groq API)",          "description": "Always uses Groq's hosted Whisper. Needs a Groq API key."},
            {"value": "local", "label": "Local Only (On-device Whisper)", "description": "Runs faster-whisper on your machine. No key needed."},
        ],
    }

@app.post("/api/stt-settings")
async def update_stt_settings(payload: STTModeSchema, db: AsyncSession = Depends(get_db)):
    if IS_PRODUCTION:
        raise HTTPException(status_code=403, detail="STT mode is locked to 'cloud' in production.")
    if payload.mode not in VALID_STT_MODES:
        raise HTTPException(status_code=400, detail=f"Invalid mode. Must be one of {sorted(VALID_STT_MODES)}")
    await set_app_setting(db, "stt_mode", payload.mode)
    os.environ["STT_MODE"] = payload.mode
    return {"status": "saved", "mode": payload.mode}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
